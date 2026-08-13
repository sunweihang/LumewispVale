#!/usr/bin/env python3
"""BOOTSTRAP ONLY — not the source of truth.

Authoritative data lives in Luban Excel: SourceData/Datas/*.xlsx
`npm run gen:config` reads those Excels directly and must NOT call this script.

Only run this manually if you need to recreate workbook shells from SLG templates;
it will overwrite Datas Excels. Prefer editing the xlsx files instead.
"""
import os
import shutil
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
DATAS = ROOT / "SourceData" / "Datas"
SLG_DATAS = Path(
    os.environ.get(
        "SLG_LUBAN_DATAS",
        "/Users/sunix/SLG/SourceData/SourceData/Datas",
    )
)


def write_data(name, sheet, header_vars, header_types, header_groups, header_comments, rows):
    shell = SLG_DATAS / "item.xlsx"
    if not shell.is_file():
        # Fall back to an existing Vale workbook shell when SLG is unavailable.
        shell = DATAS / name
        if not shell.is_file():
            raise FileNotFoundError(f"Need SLG item.xlsx or existing {name} to clone headers")
    shutil.copy(shell, DATAS / name)
    wb = openpyxl.load_workbook(DATAS / name)
    ws = wb.active
    ws.title = sheet
    ws.delete_rows(1, ws.max_row)
    ws.append(["##var"] + header_vars)
    ws.append(["##"] + [None] * len(header_vars))
    ws.append(["##type"] + header_types)
    ws.append(["##group"] + header_groups)
    ws.append(["##"] + header_comments)
    for r in rows:
        ws.append([None] + list(r))
    wb.save(DATAS / name)


def ensure_slg_shells():
    if not SLG_DATAS.is_dir():
        # Editing committed Datas in place — enums/tables still need rewriting below.
        return False
    for name in ("__enums__.xlsx", "__beans__.xlsx", "__tables__.xlsx"):
        shutil.copy(SLG_DATAS / name, DATAS / name)
    return True


def main():
    DATAS.mkdir(parents=True, exist_ok=True)
    ensure_slg_shells()

    # ── Enums ──────────────────────────────────────────────────────────────
    wb = openpyxl.load_workbook(DATAS / "__enums__.xlsx")
    ws = wb.active
    ws.delete_rows(4, ws.max_row - 3)
    for row in [
        ["ConditionOperatorType", "ConditionOperatorType", False, True, None, "条件比较操作符", None, "Less", "小于", 0, "<", None],
        [None, None, None, None, None, None, None, "Greater", "大于", 1, ">", None],
        [None, None, None, None, None, None, None, "LessEqual", "小于等于", 2, "<=", None],
        [None, None, None, None, None, None, None, "GreaterEqual", "大于等于", 3, ">=", None],
        [None, None, None, None, None, None, None, "Equal", "等于", 4, "==", None],
        [None, None, None, None, None, None, None, "NotEqual", "不等于", 5, "!=", None],
        ["ConditionType", "ConditionType", False, True, None, "条件数据源", None, "ItemCount", "背包数量", 1, "param=itemId", None],
        [None, None, None, None, None, None, None, "GatherCount", "累计采集", 2, "param=itemId", None],
        [None, None, None, None, None, None, None, "TillCount", "累计锄地", 3, "", None],
        [None, None, None, None, None, None, None, "PlantCount", "累计播种", 4, "", None],
        [None, None, None, None, None, None, None, "WaterCount", "累计浇水", 5, "", None],
        [None, None, None, None, None, None, None, "HarvestCount", "累计收获", 6, "", None],
        [None, None, None, None, None, None, None, "CraftCount", "累计合成", 7, "param=recipeId", None],
        [None, None, None, None, None, None, None, "FishCount", "累计钓鱼", 8, "", None],
        [None, None, None, None, None, None, None, "Gold", "金币数量", 9, "", None],
        [None, None, None, None, None, None, None, "Flag", "剧情旗标", 10, "param=flagId", None],
        ["GotoAction", "GotoAction", False, True, None, "引导动作", None, "None", "无动作", 0, "", None],
        [None, None, None, None, None, None, None, "OpenCraft", "打开工作台", 1, "", None],
        [None, None, None, None, None, None, None, "SelectHoe", "选中锄头", 2, "", None],
        [None, None, None, None, None, None, None, "SelectSeeds", "选中种子", 3, "", None],
        [None, None, None, None, None, None, None, "SelectCan", "选中水壶", 4, "", None],
        [None, None, None, None, None, None, None, "SelectHand", "选中手", 5, "", None],
        [None, None, None, None, None, None, None, "SelectRod", "选中鱼竿", 6, "", None],
        [None, None, None, None, None, None, None, "OpenBag", "打开背包", 7, "", None],
        [None, None, None, None, None, None, None, "HintFarm", "提示农田", 8, "", None],
        [None, None, None, None, None, None, None, "HintGrass", "提示杂草", 9, "", None],
        [None, None, None, None, None, None, None, "HintFish", "提示钓鱼", 10, "", None],
        [None, None, None, None, None, None, None, "HintCraft", "提示工作台", 11, "", None],
        [None, None, None, None, None, None, None, "HintMeteor", "提示通往小镇（旧）", 12, "", None],
        [None, None, None, None, None, None, None, "HintTownGate", "提示通往小镇的路牌", 13, "", None],
        [None, None, None, None, None, None, None, "HintMayor", "提示镇长府", 14, "", None],
        [None, None, None, None, None, None, None, "SelectAxe", "选中斧头", 15, "", None],
        [None, None, None, None, None, None, None, "HintRock", "提示挖石", 16, "", None],
        ["ItemType", "ItemType", False, True, None, "物品类型", None, "Misc", "杂项", 0, "", None],
        [None, None, None, None, None, None, None, "Tool", "工具", 1, "", None],
        [None, None, None, None, None, None, None, "Material", "材料", 2, "", None],
        [None, None, None, None, None, None, None, "Consumable", "消耗品", 3, "", None],
        [None, None, None, None, None, None, None, "Currency", "货币", 4, "", None],
        [None, None, None, None, None, None, None, "Crop", "作物", 5, "", None],
        [None, None, None, None, None, None, None, "Recipe", "配方", 6, "", None],
        [None, None, None, None, None, None, None, "Seed", "种子", 7, "", None],
    ]:
        ws.append(row)
    wb.save(DATAS / "__enums__.xlsx")

    wb = openpyxl.load_workbook(DATAS / "__beans__.xlsx")
    ws = wb.active
    ws.delete_rows(4, ws.max_row - 3)
    wb.save(DATAS / "__beans__.xlsx")

    wb = openpyxl.load_workbook(DATAS / "__tables__.xlsx")
    ws = wb.active
    ws.delete_rows(4, ws.max_row - 3)
    for row in [
        [None, "TCondition", "CCondition", True, "condition.xlsx", None, "map", "c", "条件模板", None, None],
        [None, "TCraftRecipe", "CCraftRecipe", True, "craft_recipe.xlsx", None, "map", "c", "合成配方", None, None],
        [None, "TCraftCost", "CCraftCost", True, "craft_cost.xlsx", None, "map", "c", "合成消耗", None, None],
        [None, "TQuest", "CQuest", True, "quest.xlsx", None, "map", "c", "主线任务", None, None],
        [None, "TGoto", "CGoto", True, "goto.xlsx", None, "map", "c", "引导跳转", None, None],
        [None, "TFlag", "CFlag", True, "flag.xlsx", None, "map", "c", "剧情旗标", None, None],
        [None, "TItem", "CItem", True, "item.xlsx", None, "map", "c", "基础物品表", None, None],
        [None, "TDisplay", "CDisplay", True, "display.xlsx", None, "map", "c", "展示配置表", None, None],
        [None, "TDisplayTemplate", "CDisplayTemplate", True, "display_template.xlsx", None, "map", "c", "展示模板表", None, None],
        [None, "TDialogue", "CDialogue", True, "dialogue.xlsx", None, "map", "c", "对白脚本目录", None, None],
        [None, "TChat", "CChat", True, "chat.xlsx", None, "map", "c", "对白台词行", None, None],
    ]:
        ws.append(row)
    wb.save(DATAS / "__tables__.xlsx")

    # ── Condition ──────────────────────────────────────────────────────────
    write_data(
        "condition.xlsx",
        "condition",
        ["id", "type", "compare_mode", "desc", "goto_id"],
        ["int", "ConditionType", "ConditionOperatorType", "string", "int"],
        ["c", "c", "c", "c", "c"],
        ["ID", "数据源", "比较", "描述", "跳转Goto表ID"],
        [
            [1, "GatherCount", "GreaterEqual", "采集{1} ×{0}", 9],
            [2, "TillCount", "GreaterEqual", "锄地 {0} 次", 2],
            [3, "CraftCount", "GreaterEqual", "合成「{1}」×{0}", 11],
            [4, "PlantCount", "GreaterEqual", "播种 {0} 次", 3],
            [5, "WaterCount", "GreaterEqual", "浇水 {0} 次", 4],
            [6, "HarvestCount", "GreaterEqual", "收获 {0} 次", 5],
            [7, "FishCount", "GreaterEqual", "钓到鱼 {0} 条", 6],
            [8, "ItemCount", "GreaterEqual", "背包拥有 {1} ×{0}", 7],
            [9, "Gold", "GreaterEqual", "金币达到 {0}", 0],
            [10, "Flag", "GreaterEqual", "完成：{1}", 0],
        ],
    )

    # ── Goto ───────────────────────────────────────────────────────────────
    write_data(
        "goto.xlsx",
        "goto",
        ["id", "action", "hint"],
        ["int", "GotoAction", "string"],
        ["c", "c", "c"],
        ["ID", "动作", "提示文案"],
        [
            [0, "None", ""],
            [2, "SelectHoe", "请选中锄头，点击荒地开垦"],
            [3, "SelectSeeds", "请选中种子，点击翻好地块播种"],
            [4, "SelectCan", "请选中水壶，给作物浇水"],
            [5, "SelectHand", "把催熟剂拖到快捷栏使用，再空手收获"],
            [6, "SelectRod", "先选鱼竿，跟着箭头走到西边码头，点湖面抛竿"],
            [7, "OpenBag", "打开背包查看物品"],
            [8, "HintFarm", "前往农田地块操作"],
            [9, "HintGrass", "用手拔除院子里的杂草"],
            [10, "HintFish", "跟着箭头走到湖边码头，点击湖面抛竿"],
            [11, "HintCraft", "走到工作台旁点击打开合成"],
            [12, "HintTownGate", "跟着箭头走到东侧「通往小镇」路牌，点击前往"],
            [13, "HintTownGate", "跟着箭头走到东侧「通往小镇」路牌，点击前往"],
            [14, "HintMayor", "往北走到镇长府，点大门进屋，再跟镇长·艾岚打招呼"],
            [15, "None", "前往警察局或邮局，接取公告板任务"],
            [16, "None", "前往东市木工坊，点击了解工匠"],
            [17, "None", "前往社区中心，查看修复工程"],
            [18, "None", "走进任意商店，购买一件商品"],
            [19, "None", "打开商店，点「出售」卖掉一件收获物"],
            [20, "None", "回到社区中心，在春厅收集包上签字"],
            [21, "None", "前往微光诊所，听取矿洞叮嘱"],
            [22, "None", "走进矿脉商会，向掌柜打听浅层矿洞放行"],
            [23, "None", "点击北山「通往浅层矿洞」路牌进入"],
            [24, "None", "选中锄头，点击铜矿石开采"],
            [25, "None", "带着铜矿回到社区中心，点亮春厅"],
            [26, "HintRock", "选中锄头，点击石子或岩石挖取石料"],
            [27, "SelectAxe", "选中斧头，点击松树或橡树砍伐"],
        ],
    )

    # ── Craft recipe (+ unlock gate) ───────────────────────────────────────
    write_data(
        "craft_recipe.xlsx",
        "craft_recipe",
        [
            "id",
            "name",
            "desc",
            "out_item",
            "out_count",
            "craft_seconds",
            "sort",
            "unlock_quest",
            "unlock_mode",
        ],
        ["string", "string", "string", "string", "int", "int", "int", "int", "string"],
        ["c"] * 9,
        [
            "配方ID",
            "名称",
            "描述",
            "产出物品",
            "产出数量",
            "耗时秒",
            "排序",
            "解锁任务ID",
            "解锁模式 reached|completed|空=常驻",
        ],
        [
            ["seed_from_grass", "种子", "把草料搓成可播种的种子", "seeds", 1, 12, 10, 1003, "reached"],
            ["can_basic", "水壶", "石片与泥土捏成的浇水壶", "can", 1, 16, 25, 1031, "reached"],
            ["axe_basic", "斧头", "石刃裹草绳，能砍倒松树与橡树", "axe", 1, 18, 26, 1032, "reached"],
            ["rod_basic", "鱼竿", "木杆缠草线，湖边抛竿用", "rod", 1, 18, 27, 1034, "reached"],
        ],
    )

    write_data(
        "craft_cost.xlsx",
        "craft_cost",
        ["id", "recipe_id", "item_id", "count"],
        ["int", "string", "string", "int"],
        ["c", "c", "c", "c"],
        ["ID", "配方ID", "消耗物品", "数量"],
        [
            [1, "seed_from_grass", "grass", 3],
            [6, "can_basic", "stone", 2],
            [7, "can_basic", "dirt", 3],
            [8, "axe_basic", "stone", 2],
            [9, "axe_basic", "grass", 3],
            [10, "rod_basic", "wood", 3],
            [11, "rod_basic", "grass", 2],
        ],
    )

    # ── Quest (+ intro / chapter / map unlock) ─────────────────────────────
    # chapter: farm | town | market | spring — used by GM jumps & tutorial band
    # unlock_map: when quest is active or completed, unlock that StoryMapId
    # intro_script / outro_script: StoryDialogue script ids (empty = none)
    write_data(
        "quest.xlsx",
        "quest",
        [
            "id",
            "name",
            "desc",
            "condition_id",
            "param",
            "num",
            "goto_id",
            "reward_gold",
            "reward_item",
            "reward_count",
            "next_id",
            "sort",
            "intro_script",
            "outro_script",
            "chapter",
            "unlock_map",
        ],
        [
            "int",
            "string",
            "string",
            "int",
            "string",
            "int",
            "int",
            "int",
            "string",
            "int",
            "int",
            "int",
            "string",
            "string",
            "string",
            "string",
        ],
        ["c"] * 16,
        [
            "ID",
            "标题",
            "说明",
            "条件模板",
            "条件参数",
            "目标值",
            "跳转",
            "金币奖励",
            "物品奖励",
            "物品数量",
            "下一任务",
            "排序",
            "接取对白脚本",
            "完成对白脚本",
            "章节 farm|town|market|spring",
            "解锁地图 town|mine",
        ],
        [
            [1001, "清理院子", "用手拔除院子里的杂草，多收集一些草料。", 1, "grass", 8, 9, 20, "", 0, 1002, 10, "", "", "farm", ""],
            [1002, "开垦田地", "选中锄头，多翻几块荒地，攒点泥土。", 2, "", 3, 2, 25, "", 0, 1030, 20, "quest_1002", "", "farm", ""],
            [1030, "采集石料", "选中锄头，挖开院子里的石子与岩石。", 1, "stone", 4, 26, 25, "", 0, 1031, 25, "quest_1030", "", "farm", ""],
            [1031, "打造水壶", "走到工作台，用石料与泥土制作水壶。", 3, "can_basic", 1, 11, 30, "", 0, 1003, 28, "quest_1031", "", "farm", ""],
            [1003, "搓出种子", "走到工作台，用草料合成种子。", 3, "seed_from_grass", 1, 11, 30, "seeds", 2, 1004, 30, "quest_1003", "", "farm", ""],
            [1004, "播种希望", "选中种子，在翻好的田地上多种几棵。", 4, "", 2, 3, 25, "", 0, 1005, 40, "quest_1004", "", "farm", ""],
            [1005, "浇灌作物", "选中水壶，给种下的作物浇水。完成后获得催熟剂。", 5, "", 2, 4, 25, "boost", 2, 1006, 50, "quest_1005", "", "farm", ""],
            [1006, "丰收时刻", "把催熟剂拖到快捷栏，点作物催熟，再空手收获。", 6, "", 2, 5, 40, "parsnip", 1, 1032, 60, "quest_1006", "", "farm", ""],
            [1032, "打造斧头", "走到工作台，用石料与草料制作斧头。", 3, "axe_basic", 1, 11, 35, "", 0, 1033, 65, "quest_1032", "", "farm", ""],
            [1033, "伐木练手", "选中斧头，砍伐松树或橡树，收集木料。", 1, "wood", 6, 27, 40, "", 0, 1034, 68, "quest_1033", "", "farm", ""],
            [1034, "编织鱼竿", "走到工作台，用木料与草料制作鱼竿。", 3, "rod_basic", 1, 11, 35, "", 0, 1007, 69, "quest_1034", "", "farm", ""],
            [1007, "湖边垂钓", "先选鱼竿，跟着箭头走到西边码头，点击湖面抛竿钓一条鱼。", 7, "", 1, 6, 50, "fish", 1, 1009, 70, "quest_1007", "", "farm", "town"],
            [1009, "去镇上报到", "点击东侧路牌，前往微光溪谷镇向镇长报到。", 10, "enter_town", 1, 13, 60, "", 0, 1010, 90, "quest_1009", "", "town", "town"],
            [1010, "镇长的茶", "进入镇长府，与镇长·艾岚交谈，取得定居许可。", 10, "visit_mayor", 1, 14, 80, "", 0, 1011, 100, "", "", "town", "town"],
            [1011, "镇上的声音", "到警察局或邮局的公告板接取一次委托。", 10, "accept_board", 1, 15, 50, "", 0, 1012, 110, "quest_1011", "", "town", "town"],
            [1012, "工匠的钉子", "拜访木工坊，认识负责修缮的工匠。", 10, "visit_carpenter", 1, 16, 40, "", 0, 1013, 120, "quest_1012", "", "town", "town"],
            [1013, "破旧的钟楼", "前往社区中心查看修复工程，开启下一章线索。", 10, "visit_community", 1, 17, 100, "", 0, 1020, 130, "quest_1013", "ch1_done", "town", "town"],
            [1020, "日常补给", "在镇上任意商店购买一件商品，熟悉市集买卖。", 10, "shop_buy", 1, 18, 40, "", 0, 1021, 140, "quest_1020", "", "market", "town"],
            [1021, "出手盈余", "在商店切换到「出售」，卖掉一件背包里的收获物。", 10, "shop_sell", 1, 19, 40, "", 0, 1022, 150, "quest_1021", "", "market", "town"],
            [1022, "春厅立项", "回到社区中心，在春厅收集包上签字立项。", 10, "accept_spring_pack", 1, 20, 60, "", 0, 1023, 160, "quest_1022", "", "spring", "town"],
            [1023, "诊所的叮嘱", "拜访微光诊所，听医生说说矿洞安全。", 10, "visit_clinic", 1, 21, 40, "", 0, 1024, 170, "quest_1023", "", "spring", "town"],
            [1024, "矿脉通行证", "前往矿脉商会，向掌柜·赤铜打听浅层矿洞放行。", 10, "visit_oreshop", 1, 22, 50, "", 0, 1025, 180, "quest_1024", "", "spring", "town"],
            [1025, "浅层铜脉", "点击北山矿洞路牌，进入浅层矿洞。", 10, "enter_mine", 1, 23, 40, "", 0, 1026, 190, "quest_1025", "", "spring", "mine"],
            [1026, "采一袋铜", "在矿洞里选中锄头，挖取铜矿石。", 1, "copper", 3, 24, 80, "", 0, 1027, 200, "quest_1026", "", "spring", "mine"],
            [1027, "春厅微光", "把铜矿送回社区中心，点亮春厅的第一盏灯。", 10, "light_spring_hall", 1, 25, 120, "", 0, 0, 210, "quest_1027", "ch2_done", "spring", "mine"],
        ],
    )

    # ── Flags (labels + optional map unlock when noted) ────────────────────
    write_data(
        "flag.xlsx",
        "flag",
        ["id", "label", "unlock_map"],
        ["string", "string", "string"],
        ["c", "c", "c"],
        ["旗标ID", "显示名", "触发时解锁地图"],
        [
            ["enter_town", "抵达小镇", "town"],
            ["inspect_meteor", "察看陨石", "town"],
            ["visit_mayor", "拜访镇长府", ""],
            ["shop_buy", "商店购物", ""],
            ["shop_sell", "商店出售", ""],
            ["accept_board", "接取公告板", ""],
            ["visit_carpenter", "拜访木工坊", ""],
            ["visit_community", "探访社区中心", ""],
            ["accept_spring_pack", "春厅立项签字", ""],
            ["visit_clinic", "拜访微光诊所", ""],
            ["visit_oreshop", "取得矿脉通行证", "mine"],
            ["enter_mine", "进入浅层矿洞", "mine"],
            ["light_spring_hall", "点亮春厅", ""],
        ],
    )

    # ── Items + display (SLG split). Authoritative rows live in Datas/*.xlsx. ─
    write_data(
        "item.xlsx",
        "item",
        [
            "id",
            "type",
            "name",
            "display_id",
            "use_condition_id",
            "use_effect_id",
            "max_stack",
            "desc",
            "gm_grant",
            "gm_amount",
            "sort",
        ],
        [
            "string",
            "ItemType",
            "string",
            "int",
            "int",
            "int",
            "int",
            "string",
            "bool",
            "int",
            "int",
        ],
        ["c"] * 11,
        [
            "物品ID",
            "类型",
            "显示名",
            "展示配置ID",
            "使用条件ID",
            "使用效果ID",
            "最大堆叠",
            "描述",
            "GM可发放",
            "GM单次数量",
            "排序",
        ],
        [
            ["hoe", "Tool", "锄头", 1002, 0, 0, 1, "开垦荒地", True, 1, 20],
        ],
    )
    write_data(
        "display_template.xlsx",
        "display_template",
        ["id", "name", "desc", "param_desc"],
        ["int", "string", "string", "string"],
        ["c", "c", "c", "c"],
        ["ID", "模板名称", "说明", "参数说明"],
        [
            [1, "堆叠数量", "显示道具数量", "num=最大显示数量"],
            [11, "图标展示", "显示道具图标与分类名", "param JSON: icon, kind"],
            [12, "售价", "可出售标价（无此行=不可出售）", "param JSON: price, currency"],
        ],
    )
    write_data(
        "display.xlsx",
        "display",
        ["id", "display_template_id", "param", "num", "link_id"],
        ["int", "int", "string", "int", "int"],
        ["c", "c", "c", "c", "c"],
        ["ID", "展示模板ID", "参数(JSON)", "数值", "关联物品display_id"],
        [
            [1, 11, '{"icon":"","kind":"工具"}', 0, 1002],
            [2, 1, "[]", 1, 1002],
        ],
    )

    conf = ROOT / "SourceData" / "luban.conf"
    conf.write_text(
        "{\n"
        '\t"groups":\n'
        "\t[\n"
        '\t\t{"names":["c"], "default":true},\n'
        '\t\t{"names":["s"], "default":true},\n'
        '\t\t{"names":["e"], "default":true}\n'
        "\t],\n"
        '\t"schemaFiles":\n'
        "\t[\n"
        '\t\t{"fileName":"Datas/__tables__.xlsx", "type":"table"},\n'
        '\t\t{"fileName":"Datas/__beans__.xlsx", "type":"bean"},\n'
        '\t\t{"fileName":"Datas/__enums__.xlsx", "type":"enum"}\n'
        "\t],\n"
        '\t"dataDir": "Datas",\n'
        '\t"targets":\n'
        "\t[\n"
        '\t\t{"name":"client", "manager":"Tables", "groups":["c"], "topModule":"cfg"}\n'
        "\t]\n"
        "}\n",
        encoding="utf-8",
    )
    print("Wrote Luban source under", DATAS)


if __name__ == "__main__":
    main()
