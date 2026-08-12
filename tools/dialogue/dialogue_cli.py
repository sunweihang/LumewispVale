#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Lumewisp Vale — Luban dialogue CLI (token-cheap path for agents).

  python3 tools/dialogue/dialogue_cli.py status
  python3 tools/dialogue/dialogue_cli.py show wake_farm
  python3 tools/dialogue/dialogue_cli.py add --json '{"script_id":"x","name":"名","lines":[{"speaker":"露穗","text":"…"}]}'
  python3 tools/dialogue/dialogue_cli.py apply
  python3 tools/dialogue/dialogue_cli.py verify

Do NOT hand-edit tdialogue.json / tchat.json / DialogueScripts.ts.
"""
from __future__ import print_function

import argparse
import json
import os
import re
import subprocess
import sys

try:
    import openpyxl
except ImportError:
    print("ERROR: need openpyxl (pip install openpyxl)", file=sys.stderr)
    sys.exit(2)

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "../.."))
DIALOGUE_XLSX = os.path.join(ROOT, "SourceData/Datas/dialogue.xlsx")
CHAT_XLSX = os.path.join(ROOT, "SourceData/Datas/chat.xlsx")
QUEST_XLSX = os.path.join(ROOT, "SourceData/Datas/quest.xlsx")
TDIALOGUE_JSON = os.path.join(ROOT, "assets/resources/config/tdialogue.json")
TCHAT_JSON = os.path.join(ROOT, "assets/resources/config/tchat.json")
TQUEST_JSON = os.path.join(ROOT, "assets/resources/config/tquest.json")

SCRIPT_ID_RE = re.compile(r"^[a-z][a-z0-9_]*$")


def _load_sheet(path, sheet=None):
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet] if sheet else wb.active
    return wb, ws


def _data_rows(ws):
    """Yield (excel_row_index, values_tuple) for rows after header (row>=6)."""
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i <= 5:
            continue
        if row[1] is None:
            continue
        yield i, row


def _read_dialogues():
    """script_id -> {id, script_id, kind, name}"""
    _, ws = _load_sheet(DIALOGUE_XLSX)
    out = {}
    for _, row in _data_rows(ws):
        # cols: A empty, B id, C script_id, D kind, E name
        did = int(row[1])
        sid = str(row[2]).strip()
        out[sid] = {
            "id": did,
            "script_id": sid,
            "kind": str(row[3] or "dialogue").strip(),
            "name": str(row[4] or sid).strip(),
        }
    return out


def _read_chats():
    """dialogue_id -> list of line dicts"""
    _, ws = _load_sheet(CHAT_XLSX)
    by = {}
    max_id = 0
    for _, row in _data_rows(ws):
        lid = int(row[1])
        did = int(row[2])
        max_id = max(max_id, lid)
        by.setdefault(did, []).append(
            {
                "id": lid,
                "dialogue_id": did,
                "seq": int(row[3]),
                "speaker": str(row[4] or ""),
                "text": str(row[5] or ""),
                "image": str(row[6] or ""),
            }
        )
    for lst in by.values():
        lst.sort(key=lambda x: (x["seq"], x["id"]))
    return by, max_id


def _next_dialogue_id(dialogues):
    if not dialogues:
        return 10001
    return max(d["id"] for d in dialogues.values()) + 1


def cmd_status(_args):
    d = _read_dialogues()
    chats, max_line = _read_chats()
    nxt = _next_dialogue_id(d)
    print("dialogues\t%d" % len(d))
    print("chat_lines\t%d" % sum(len(v) for v in chats.values()))
    print("next_dialogue_id\t%d" % nxt)
    print("next_chat_id\t%d" % (max_line + 1))
    print("max_dialogue_id\t%d" % (nxt - 1 if d else 0))


def cmd_list(_args):
    d = _read_dialogues()
    chats, _ = _read_chats()
    for sid, meta in sorted(d.items(), key=lambda x: x[1]["id"]):
        n = len(chats.get(meta["id"], []))
        print("%d\t%s\t%s\t%d" % (meta["id"], sid, meta["kind"], n))


def cmd_show(args):
    key = args.key
    d = _read_dialogues()
    chats, _ = _read_chats()
    meta = None
    if key.isdigit():
        did = int(key)
        meta = next((m for m in d.values() if m["id"] == did), None)
    else:
        meta = d.get(key)
    if not meta:
        print("ERROR: not found: %s" % key, file=sys.stderr)
        sys.exit(1)
    print("id\t%d" % meta["id"])
    print("script_id\t%s" % meta["script_id"])
    print("kind\t%s" % meta["kind"])
    print("name\t%s" % meta["name"])
    for line in chats.get(meta["id"], []):
        sp = line["speaker"] or "(旁白)"
        img = "\timage=%s" % line["image"] if line["image"] else ""
        print("%d\t%s\t%s%s" % (line["seq"], sp, line["text"], img))


def _parse_payload(args):
    if args.json:
        raw = args.json
    elif args.file:
        with open(args.file, "r", encoding="utf-8") as f:
            raw = f.read()
    elif not sys.stdin.isatty():
        raw = sys.stdin.read()
    else:
        print("ERROR: pass --json, --file, or stdin JSON", file=sys.stderr)
        sys.exit(2)
    try:
        data = json.loads(raw)
    except ValueError as e:
        print("ERROR: invalid JSON: %s" % e, file=sys.stderr)
        sys.exit(2)
    return data


def _validate_add(data, dialogues):
    sid = (data.get("script_id") or "").strip()
    if not sid or not SCRIPT_ID_RE.match(sid):
        raise ValueError("script_id must match [a-z][a-z0-9_]*")
    if sid in dialogues and not data.get("replace"):
        raise ValueError("script_id already exists: %s (pass replace:true to overwrite lines)" % sid)
    kind = (data.get("kind") or "dialogue").strip()
    if kind not in ("dialogue", "intro"):
        raise ValueError("kind must be dialogue|intro")
    lines = data.get("lines")
    if not isinstance(lines, list) or not lines:
        raise ValueError("lines must be a non-empty array")
    for i, line in enumerate(lines):
        if not isinstance(line, dict):
            raise ValueError("lines[%d] must be object" % i)
        if not (line.get("text") or "").strip():
            raise ValueError("lines[%d].text required" % i)
        if kind == "intro" and not (line.get("image") or "").strip():
            raise ValueError("intro lines[%d].image (SpriteFrame uuid) required" % i)
    return sid, kind, lines


def _set_quest_script(quest_id, field, script_id):
    """field: intro_script | outro_script"""
    wb, ws = _load_sheet(QUEST_XLSX)
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    # headers[0]=##var, then id at 1, ...
    try:
        id_col = headers.index("id")
        field_col = headers.index(field)
    except ValueError:
        raise ValueError("quest.xlsx missing column %s" % field)
    found = False
    for i, row in enumerate(ws.iter_rows(min_row=6), 6):
        cell_id = row[id_col].value
        if cell_id is None:
            continue
        if int(cell_id) == int(quest_id):
            row[field_col].value = script_id
            found = True
            break
    if not found:
        raise ValueError("quest id %s not found" % quest_id)
    wb.save(QUEST_XLSX)


def cmd_add(args):
    data = _parse_payload(args)
    dialogues = _read_dialogues()
    sid, kind, lines = _validate_add(data, dialogues)
    name = (data.get("name") or sid).strip()
    replace = bool(data.get("replace"))

    chats, max_line = _read_chats()
    next_line = max_line + 1

    if sid in dialogues:
        did = dialogues[sid]["id"]
        if not replace:
            raise ValueError("exists")
        # wipe chat rows for this dialogue_id
        wb_c, ws_c = _load_sheet(CHAT_XLSX)
        to_delete = []
        for excel_i, row in _data_rows(ws_c):
            if int(row[2]) == did:
                to_delete.append(excel_i)
        for excel_i in reversed(to_delete):
            ws_c.delete_rows(excel_i, 1)
        # update dialogue meta
        wb_d, ws_d = _load_sheet(DIALOGUE_XLSX)
        for excel_i, row in _data_rows(ws_d):
            if str(row[2]).strip() == sid:
                ws_d.cell(excel_i, 4).value = kind  # D kind
                ws_d.cell(excel_i, 5).value = name  # E name
                break
        wb_d.save(DIALOGUE_XLSX)
        # re-open chat after delete
        wb_c.save(CHAT_XLSX)
        wb_c, ws_c = _load_sheet(CHAT_XLSX)
    else:
        did = int(data["id"]) if data.get("id") else _next_dialogue_id(dialogues)
        if any(m["id"] == did for m in dialogues.values()):
            raise ValueError("dialogue id %d already used" % did)
        wb_d, ws_d = _load_sheet(DIALOGUE_XLSX)
        ws_d.append([None, did, sid, kind, name])
        wb_d.save(DIALOGUE_XLSX)
        wb_c, ws_c = _load_sheet(CHAT_XLSX)

    for seq, line in enumerate(lines, 1):
        speaker = (line.get("speaker") or "").strip()
        text = (line.get("text") or "").strip()
        image = (line.get("image") or "").strip()
        ws_c.append([None, next_line, did, seq, speaker, text, image])
        next_line += 1
    wb_c.save(CHAT_XLSX)

    notes = []
    if data.get("quest_intro") is not None:
        _set_quest_script(int(data["quest_intro"]), "intro_script", sid)
        notes.append("quest %s intro_script=%s" % (data["quest_intro"], sid))
    if data.get("quest_outro") is not None:
        _set_quest_script(int(data["quest_outro"]), "outro_script", sid)
        notes.append("quest %s outro_script=%s" % (data["quest_outro"], sid))

    print("OK\tid=%d\tscript_id=%s\tkind=%s\tlines=%d" % (did, sid, kind, len(lines)))
    for n in notes:
        print("WIRE\t%s" % n)
    print("NEXT\trun: python3 tools/dialogue/dialogue_cli.py apply")


def cmd_apply(_args):
    print(">> npm run gen:config")
    r = subprocess.call(["npm", "run", "gen:config"], cwd=ROOT)
    if r != 0:
        sys.exit(r)
    print(">> node tools/story/generate_dialogue_stories.mjs")
    r = subprocess.call(["node", "tools/story/generate_dialogue_stories.mjs"], cwd=ROOT)
    if r != 0:
        sys.exit(r)
    print("OK\tconfig+story-graphs refreshed")


def cmd_verify(_args):
    errors = []
    # Prefer generated JSON if present; else Excel
    if os.path.isfile(TDIALOGUE_JSON) and os.path.isfile(TCHAT_JSON):
        dialogues = {r["script_id"]: r for r in json.load(open(TDIALOGUE_JSON))}
        chats = json.load(open(TCHAT_JSON))
        by = {}
        for r in chats:
            by.setdefault(r["dialogue_id"], []).append(r)
    else:
        dialogues = _read_dialogues()
        by, _ = _read_chats()
        # normalize keys
        dialogues = {k: {"id": v["id"], "script_id": k, "kind": v["kind"]} for k, v in dialogues.items()}

    for sid, meta in dialogues.items():
        did = meta["id"]
        lines = by.get(did, [])
        if not lines:
            errors.append("dialogue %d (%s) has no chat lines" % (did, sid))
        if meta.get("kind") == "intro":
            for ln in lines:
                img = ln.get("image") if isinstance(ln, dict) else ""
                if not img:
                    errors.append("intro %s missing image on a line" % sid)

    if os.path.isfile(TQUEST_JSON):
        for q in json.load(open(TQUEST_JSON)):
            for field in ("intro_script", "outro_script"):
                sid = (q.get(field) or "").strip()
                if sid and sid not in dialogues:
                    errors.append("quest %s %s=%s missing in TDialogue" % (q.get("id"), field, sid))

    # story graph presence for each dialogue id
    for sid, meta in dialogues.items():
        did = meta["id"]
        g = os.path.join(ROOT, "assets/resources/story-graphs", str(did), "graph.graph.json")
        ts = os.path.join(ROOT, "assets/scripts/story/generated", "TsStory%d.ts" % did)
        if not os.path.isfile(g):
            errors.append("missing story graph %d" % did)
        if not os.path.isfile(ts):
            errors.append("missing TsStory%d.ts" % did)

    if errors:
        print("FAIL\t%d" % len(errors))
        for e in errors:
            print("ERR\t%s" % e)
        sys.exit(1)
    print("OK\tdialogues=%d" % len(dialogues))


def cmd_template(_args):
    """Print a minimal JSON template (copy-fill)."""
    print(
        json.dumps(
            {
                "script_id": "quest_XXXX",
                "name": "任务XXXX",
                "kind": "dialogue",
                "lines": [
                    {"speaker": "露穗", "text": "第一句"},
                    {"speaker": "", "text": "旁白句（speaker 空）"},
                ],
                "quest_intro": 0,
            },
            ensure_ascii=False,
            indent=2,
        )
    )


def main():
    ap = argparse.ArgumentParser(description="Luban dialogue CLI")
    sub = ap.add_subparsers(dest="cmd")

    sub.add_parser("status", help="next ids + counts")
    sub.add_parser("list", help="list script_id")
    p_show = sub.add_parser("show", help="show one script")
    p_show.add_argument("key", help="script_id or numeric id")

    p_add = sub.add_parser("add", help="append dialogue+lines to Excel")
    p_add.add_argument("--json", help="JSON payload string")
    p_add.add_argument("--file", help="JSON file path")

    sub.add_parser("apply", help="gen:config + regenerate story graphs")
    sub.add_parser("verify", help="integrity check")
    sub.add_parser("template", help="print JSON template")

    args = ap.parse_args()
    if not args.cmd:
        ap.print_help()
        sys.exit(2)
    try:
        if args.cmd == "status":
            cmd_status(args)
        elif args.cmd == "list":
            cmd_list(args)
        elif args.cmd == "show":
            cmd_show(args)
        elif args.cmd == "add":
            cmd_add(args)
        elif args.cmd == "apply":
            cmd_apply(args)
        elif args.cmd == "verify":
            cmd_verify(args)
        elif args.cmd == "template":
            cmd_template(args)
    except ValueError as e:
        print("ERROR: %s" % e, file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
